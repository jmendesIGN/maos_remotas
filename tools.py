"""
Ferramentas disponíveis ao agente GPT-4.1.
Cada função aqui corresponde a uma "tool" que o agente pode chamar.
"""

import json
import logging
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path

import config

log = logging.getLogger(__name__)

# ── Cores das células ──────────────────────────────────────────────────────────
COR_VERDE    = PatternFill(fill_type="solid", fgColor="92D050")  # SID encontrado
COR_VERMELHO = PatternFill(fill_type="solid", fgColor="FF6B6B")  # Não encontrado
COR_AMARELO  = PatternFill(fill_type="solid", fgColor="FFD966")  # Caso especial

# ── Estado global da planilha (compartilhado entre chamadas de ferramenta) ─────
_workbook: openpyxl.Workbook | None = None
_worksheet = None
_caminho_arquivo: Path | None = None
_col_sid: int = 15  # Coluna SID (padrão: coluna 15)


# ── Inicialização da planilha ─────────────────────────────────────────────────
def inicializar_planilha(caminho: str) -> list[dict]:
    """
    Carrega a planilha e retorna as linhas como lista de dicionários.
    Chamado pelo app.py antes de iniciar o agente.
    """
    global _workbook, _worksheet, _caminho_arquivo, _col_sid

    path = Path(caminho)
    _caminho_arquivo = path.with_suffix(".xlsx")

    if path.suffix.lower() == ".csv":
        _workbook = _csv_para_workbook(path)
    else:
        _workbook = openpyxl.load_workbook(path)

    _worksheet = _workbook.active

    # Descobre a coluna SID pelo cabeçalho
    cabecalho = [str(c.value or "").strip().upper() for c in _worksheet[1]]
    if "SID" in cabecalho:
        _col_sid = cabecalho.index("SID") + 1

    # Converte linhas para lista de dicionários
    nomes_colunas = [str(c.value or "").strip() for c in _worksheet[1]]
    linhas = []
    for i, row in enumerate(_worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        linha = {"_numero_linha": i}
        for col_nome, valor in zip(nomes_colunas, row):
            linha[col_nome] = str(valor or "").strip()
        linhas.append(linha)

    log.info(f"Planilha carregada: {len(linhas)} linhas, coluna SID = {_col_sid}")
    return linhas


def _csv_para_workbook(path: Path) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(path, encoding="utf-8-sig") as f:
        for linha in f:
            ws.append(linha.rstrip("\n").split(";"))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return wb


# ── Definições das ferramentas (schema para o agente) ─────────────────────────
TOOLS_SCHEMA = [
    {
        "type": "function",
        "function": {
            "name": "buscar_no_quickbase",
            "description": (
                "Busca um serviço no Quickbase usando o identificador extraído da coluna "
                "IGN Network da planilha Skynet. Retorna o SID do serviço se encontrado, "
                "ou null se não existir."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "tipo": {
                        "type": "string",
                        "enum": ["PID", "TT", "ACC", "TECH_ACTIVITY"],
                        "description": "Tipo do identificador extraído da coluna IGN Network.",
                    },
                    "valor": {
                        "type": "string",
                        "description": (
                            "Valor do identificador. Ex: '15241' para PID15241, "
                            "'293815' para TT293815, 'ACC-4242771-34437-13' para ACC."
                        ),
                    },
                },
                "required": ["tipo", "valor"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "atualizar_sid_planilha",
            "description": (
                "Preenche a coluna SID de uma linha específica da planilha com o valor "
                "retornado pelo Quickbase. Aplica formatação de cor: verde (encontrado), "
                "vermelho (NULL), amarelo (caso especial PID=SID)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "numero_linha": {
                        "type": "integer",
                        "description": "Número da linha na planilha (campo _numero_linha).",
                    },
                    "sid": {
                        "type": "string",
                        "description": "Valor do SID retornado pelo Quickbase, ou 'NULL' se não encontrado.",
                    },
                    "encontrado": {
                        "type": "boolean",
                        "description": "True se o serviço foi encontrado no Quickbase.",
                    },
                    "especial": {
                        "type": "boolean",
                        "description": "True se o SID retornado é igual ao próprio identificador (caso PID=SID).",
                    },
                },
                "required": ["numero_linha", "sid", "encontrado"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "atualizar_status_quickbase",
            "description": (
                "Atualiza o campo status de um registro no Quickbase para 'Pago'. "
                "Deve ser chamado apenas para serviços encontrados e com SID válido."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "tipo": {
                        "type": "string",
                        "enum": ["PID", "TT", "ACC", "TECH_ACTIVITY"],
                        "description": "Tipo do identificador.",
                    },
                    "valor": {
                        "type": "string",
                        "description": "Valor do identificador para localizar o registro.",
                    },
                },
                "required": ["tipo", "valor"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "salvar_planilha",
            "description": "Salva todas as alterações feitas na planilha no arquivo original.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
]


# ── Implementações das ferramentas ─────────────────────────────────────────────
def buscar_no_quickbase(tipo: str, valor: str) -> dict:
    """Consulta o Quickbase e retorna o SID ou null."""
    field_map = {
        "PID":           config.FIELD_ID_PID,
        "TT":            config.FIELD_ID_TT,
        "ACC":           config.FIELD_ID_ACC,
        "TECH_ACTIVITY": config.FIELD_ID_TECH_ACTIVITY,
    }
    field_id = field_map.get(tipo, 0)

    if not field_id:
        msg = f"Field ID não configurado para o tipo '{tipo}'. Configure config.py."
        log.warning(msg)
        return {"sid": None, "erro": msg}

    headers = {
        "QB-Realm-Hostname": f"{config.QUICKBASE_REALM}.quickbase.com",
        "Authorization": f"QB-USER-TOKEN {config.QUICKBASE_USER_TOKEN}",
        "Content-Type": "application/json",
    }
    payload = {
        "from": config.QUICKBASE_TABLE_ID,
        "select": [config.FIELD_ID_SID],
        "where": f"{{{field_id}.EX.'{valor}'}}",
        "top": 1,
    }

    try:
        r = requests.post(
            "https://api.quickbase.com/v1/records/query",
            headers=headers,
            json=payload,
            timeout=15,
        )
        r.raise_for_status()
        dados = r.json()
        registros = dados.get("data", [])
        if not registros:
            log.info(f"  Quickbase: {tipo}={valor} → NÃO ENCONTRADO")
            return {"sid": None}
        sid = registros[0].get(str(config.FIELD_ID_SID), {}).get("value")
        log.info(f"  Quickbase: {tipo}={valor} → SID={sid}")
        return {"sid": str(sid) if sid is not None else None}
    except Exception as e:
        log.error(f"Erro ao consultar Quickbase ({tipo}={valor}): {e}")
        return {"sid": None, "erro": str(e)}


def atualizar_sid_planilha(numero_linha: int, sid: str, encontrado: bool, especial: bool = False) -> dict:
    """Escreve o SID na célula da planilha com a cor correta."""
    if _worksheet is None:
        return {"ok": False, "erro": "Planilha não inicializada."}

    celula = _worksheet.cell(row=numero_linha, column=_col_sid)
    celula.value = sid

    if not encontrado:
        celula.fill = COR_VERMELHO
        cor = "vermelho"
    elif especial:
        celula.fill = COR_AMARELO
        cor = "amarelo"
    else:
        celula.fill = COR_VERDE
        cor = "verde"

    log.info(f"  Linha {numero_linha}: SID='{sid}' | cor={cor}")
    return {"ok": True, "linha": numero_linha, "sid": sid, "cor": cor}


def atualizar_status_quickbase(tipo: str, valor: str) -> dict:
    """Atualiza o status do registro no Quickbase para 'Pago'."""
    if not config.FIELD_ID_STATUS:
        return {"ok": False, "erro": "FIELD_ID_STATUS não configurado em config.py."}

    field_map = {
        "PID":           config.FIELD_ID_PID,
        "TT":            config.FIELD_ID_TT,
        "ACC":           config.FIELD_ID_ACC,
        "TECH_ACTIVITY": config.FIELD_ID_TECH_ACTIVITY,
    }
    field_busca = field_map.get(tipo, 0)
    if not field_busca:
        return {"ok": False, "erro": f"Field ID não configurado para '{tipo}'."}

    headers = {
        "QB-Realm-Hostname": f"{config.QUICKBASE_REALM}.quickbase.com",
        "Authorization": f"QB-USER-TOKEN {config.QUICKBASE_USER_TOKEN}",
        "Content-Type": "application/json",
    }

    # Busca o record ID para poder atualizar
    payload_busca = {
        "from": config.QUICKBASE_TABLE_ID,
        "select": [3],  # Field 3 = Record ID# padrão no Quickbase
        "where": f"{{{field_busca}.EX.'{valor}'}}",
        "top": 1,
    }
    try:
        r = requests.post(
            "https://api.quickbase.com/v1/records/query",
            headers=headers,
            json=payload_busca,
            timeout=15,
        )
        r.raise_for_status()
        registros = r.json().get("data", [])
        if not registros:
            return {"ok": False, "erro": "Registro não encontrado para atualização de status."}

        record_id = registros[0].get("3", {}).get("value")

        payload_update = {
            "to": config.QUICKBASE_TABLE_ID,
            "data": [{
                "3": {"value": record_id},
                str(config.FIELD_ID_STATUS): {"value": "Pago"},
            }],
        }
        r2 = requests.post(
            "https://api.quickbase.com/v1/records",
            headers=headers,
            json=payload_update,
            timeout=15,
        )
        r2.raise_for_status()
        log.info(f"  Status atualizado para 'Pago': {tipo}={valor}")
        return {"ok": True, "record_id": record_id}
    except Exception as e:
        log.error(f"Erro ao atualizar status ({tipo}={valor}): {e}")
        return {"ok": False, "erro": str(e)}


def salvar_planilha() -> dict:
    """Salva o workbook no arquivo .xlsx."""
    if _workbook is None or _caminho_arquivo is None:
        return {"ok": False, "erro": "Planilha não inicializada."}
    _workbook.save(_caminho_arquivo)
    log.info(f"Planilha salva: {_caminho_arquivo}")
    return {"ok": True, "arquivo": str(_caminho_arquivo)}


# ── Dispatcher — executa a ferramenta pelo nome ────────────────────────────────
def executar_ferramenta(nome: str, argumentos: str) -> str:
    """Recebe o nome e argumentos JSON da ferramenta e retorna o resultado como string JSON."""
    args = json.loads(argumentos)
    if nome == "buscar_no_quickbase":
        resultado = buscar_no_quickbase(**args)
    elif nome == "atualizar_sid_planilha":
        resultado = atualizar_sid_planilha(**args)
    elif nome == "atualizar_status_quickbase":
        resultado = atualizar_status_quickbase(**args)
    elif nome == "salvar_planilha":
        resultado = salvar_planilha()
    else:
        resultado = {"erro": f"Ferramenta desconhecida: {nome}"}
    return json.dumps(resultado, ensure_ascii=False)
